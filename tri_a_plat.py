import streamlit as st
import pandas as pd
from scipy.stats import chi2_contingency
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def replace_apostrophe_in_columns(df, columns):
    """
    Remplace ’ par ' dans les colonnes spécifiées d'un DataFrame.

    :param df: DataFrame pandas
    :param columns: Liste des colonnes où effectuer le remplacement
    :return: DataFrame modifié
    """
    for column in columns:
        if column in df.columns:
            df[column] = df[column].astype(str).str.replace('’', "'", regex=False)
        else:
            print(f"Attention : La colonne '{column}' n'existe pas dans le DataFrame.")
    return df




def tri_a_plat(df, column, download_to_xlsx=False, title="", col_poids=None, QCM=False):
    """
    Calcule le nombre de réponses et le pourcentage pour chaque réponse dans une colonne donnée,
    où chaque ligne peut contenir des réponses multiples séparées par des virgules ou des reponses uniques.

    Parameters:
    - df (pd.DataFrame): Le DataFrame contenant les données.
    - column (str): Le nom de la colonne où chaque ligne correspond à une ou plusieurs réponses.
    - download_to_xlsx (bool): Si True, télécharge le résultat dans un fichier Excel.
    - title (str): Titre du fichier Excel (utilisé si download_to_xlsx est True).
    - col_poids (str): Le nom de la colonne contenant les poids (pour une pondération des résultats).
    - QCM (bool) : True si la question est à choix multiple, False sinon.

    Returns:
    - pd.DataFrame: Un tableau contenant les réponses, le nombre de réponses et leur pourcentage.
    """

    df = replace_apostrophe_in_columns(df, [column])
    # Vérifier si la colonne existe dans le DataFrame
    if column not in df.columns:
        raise ValueError(f"La colonne '{column}' n'existe pas dans le DataFrame.")


    df_non_vide = df[df[column] != ""]
    df_non_vide = df_non_vide[df_non_vide[column] != "nan"]
    total_respondents = df_non_vide.shape[0]

    # Diviser les réponses multiples en lignes distinctes
    if QCM:
      all_responses = df_non_vide[column].dropna().astype(str).str.split(',').explode().str.strip()
    else:
      all_responses = df_non_vide[column].dropna().astype(str)

    # Si des poids sont fournis, effectuer un calcul pondéré

    #     all_responses['Poids'] = all_responses.index.map(weights.to_dict())


    # Calculer le nombre de réponses simples
    response_counts = all_responses.value_counts()
    response_percentages_raw = (response_counts / total_respondents ).round(4)

    response_percentages = response_percentages_raw

    if col_poids:
      if col_poids not in df.columns:
          raise ValueError(f"La colonne de poids '{col_poids}' n'existe pas dans le DataFrame.")

      df_filtered = df_non_vide.copy()
      df_filtered[col_poids] = df_filtered[col_poids].astype(str).str.replace(',', '.')
      df_filtered[col_poids] = pd.to_numeric(df_filtered[col_poids], errors='coerce')

      if QCM:
        df_filtered = df_filtered[[column, col_poids]].dropna()
        df_filtered[column] = df_filtered[column].astype(str).str.split(',')
        df_filtered = df_filtered.explode(column)
        df_filtered[column] = df_filtered[column].str.strip()
      else:
        df_filtered = df_filtered[[column, col_poids]].dropna().astype(str)

      df_filtered[col_poids] = pd.to_numeric(df_filtered[col_poids], errors='coerce')


      response_counts_poids = df_filtered.groupby(column)[col_poids].sum().round(1)

      total_weight = response_counts_poids.sum()
      response_percentages_poids_raw = (response_counts_poids / total_weight)

      response_percentages_poids = response_percentages_poids_raw

      # Créer le DataFrame des résultats
      # result_df = pd.DataFrame({
      #     column: response_counts_poids.index,
      #     'Effectif': response_counts.values,
      #     'Fréquence': response_percentages.values,
      #     'Effectif pondéré': response_counts_poids.values,
      #     'Fréquence pondérée': response_percentages_poids.values
      # })

      # Aligner les séries sur les mêmes index
      all_index = response_counts_poids.index.union(response_counts.index)
      response_counts_aligned = response_counts.reindex(all_index, fill_value=0)
      response_percentages_aligned = response_percentages.reindex(all_index, fill_value=0)
      response_counts_poids_aligned = response_counts_poids.reindex(all_index, fill_value=0)
      response_percentages_poids_aligned = response_percentages_poids.reindex(all_index, fill_value=0)

      result_df = pd.DataFrame({
          column: all_index,
          'Effectif': response_counts_aligned.values,
          'Fréquence': response_percentages_aligned.values,
          'Effectif pondéré': response_counts_poids_aligned.values,
          'Fréquence pondérée': response_percentages_poids_aligned.values
      })

      # Ajouter une ligne Total
      total_effectif = response_counts.sum()
      total_frequence = response_percentages_raw.sum()
      total_effectif_poids = response_counts_poids.sum()
      total_frequence_poids = response_percentages_poids_raw.sum()
      result_df.loc[len(result_df)] = ["Total", total_effectif, total_frequence, total_effectif_poids, total_frequence_poids]

    else:
      # Créer le DataFrame des résultats
      result_df = pd.DataFrame({
          column: response_counts.index,
          'Effectif': response_counts.values,
          'Fréquence': response_percentages.values,
      })

      # Ajouter une ligne Total
      total_effectif = response_counts.sum()
      total_frequence = response_percentages_raw.sum()
      result_df.loc[len(result_df)] = ["Total", total_effectif, total_frequence]


    return result_df

def tableaux_croises_multi(df, variables, max_modalites=10):
    """
    Crée des tableaux croisés (effectifs et pourcentages) pour toutes les paires de variables.
    Ignore les variables ayant plus de max_modalites modalités.

    Args:
        df (pd.DataFrame) : dataframe brut (1 ligne = 1 individu)
        variables (list) : liste des colonnes à utiliser comme variables
        max_modalites (int) : nombre maximum de modalités autorisées pour une variable

    Returns:
        dict : {(var_row, var_col): (df_counts, df_perc)}
    """
    result = {}

    # Filtrer les variables avec trop de modalités
    variables_filtrees = [var for var in variables if df[var].nunique() <= max_modalites]

    # Toutes les combinaisons possibles de 2 variables filtrées
    for var_row in variables_filtrees:
      for var_col in variables_filtrees:
        df = df[(df[var_row]!="nan") & (df[var_col]!="nan")]
        # Préfixer les modalités avec le nom de la variable
        row_labels = df[var_row].apply(lambda x: f"{var_row} : {x}")
        col_labels = df[var_col].apply(lambda x: f"{var_col} : {x}")

        # DataFrame temporaire
        temp_df = pd.DataFrame({
            'row_label': row_labels,
            'col_label': col_labels
        })

        # Tableau des effectifs
        df_counts = pd.crosstab(temp_df['row_label'], temp_df['col_label'])
        df_perc = df_counts.div(df_counts.sum(axis=0), axis=1) * 100

        # Stocker dans le dictionnaire
        result[(var_row, var_col)] = (df_counts, df_perc)

    return result



def couleurs_khi2(resultats, output_filename="tableaux_khi2_colores.xlsx"):
    """
    Crée un tableau unique avec toutes les modalités en ligne et colonne,
    affiche les pourcentages et colorie les sur/sous-représentations basées sur chi².
    Ajoute le nom de la variable avant chaque bloc et une ligne vide après le bloc.

    Args:
        resultats (dict) : {(var_row, var_col): (df_counts, df_perc)}
        output_filename (str) : nom du fichier Excel à générer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Analyse Khi-2"

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # 1️⃣ Récupérer toutes les modalités dans l'ordre du dictionnaire
    toutes_modalites = []
    seen_modalites = set()

    for (var_row, var_col), (df_counts, _) in resultats.items():
        for m in df_counts.index:
            if m not in seen_modalites:
                toutes_modalites.append(m)
                seen_modalites.add(m)
        for m in df_counts.columns:
            if m not in seen_modalites:
                toutes_modalites.append(m)
                seen_modalites.add(m)

    # 2️⃣ Créer un grand tableau vide pour Excel
    tableau_perc = pd.DataFrame(index=toutes_modalites, columns=toutes_modalites, dtype=float)
    tableau_z = pd.DataFrame(index=toutes_modalites, columns=toutes_modalites, dtype=float)

    # 3️⃣ Remplir le tableau avec les données de chaque croisement
    for (var_row, var_col), (df_counts, df_perc) in resultats.items():
        observed = df_counts.to_numpy()
        try:
            chi2, p, dof, expected = chi2_contingency(observed)
            residus_std = (observed - expected) / np.sqrt(expected)
        except Exception as e:
            print(f"⛔ Problème sur le tableau {var_row} x {var_col}: {e}")
            continue

        rows = df_counts.index
        cols = df_counts.columns

        for i, row in enumerate(rows):
            for j, col in enumerate(cols):
                tableau_perc.loc[row, col] = df_perc.loc[row, col]
                tableau_z.loc[row, col] = residus_std[i, j] if p < 0.05 else 0

    # 4️⃣ Écrire dans Excel
    # En-tête
    ws.cell(row=1, column=1).value = "Modalité"
    for j, col in enumerate(toutes_modalites):
        ws.cell(row=1, column=j + 2).value = col

    current_var = None
    excel_row = 2  # ligne à écrire sous l'en-tête

    for row in toutes_modalites:
        var_name = row.split(" : ")[0]

        # Détecter changement de variable
        if current_var is None or var_name != current_var:
            # Ligne avec le nom de la variable
            ws.cell(row=excel_row, column=1).value = var_name
            excel_row += 1
            current_var = var_name

        # Écrire le libellé de la modalité
        ws.cell(row=excel_row, column=1).value = row

        # Remplir les cellules de la ligne
        for j, col in enumerate(toutes_modalites):
            val = tableau_perc.loc[row, col]
            if pd.isna(val):
                continue

            cell = ws.cell(row=excel_row, column=j + 2)
            cell.value = f"{val:.2f}%"

            # Griser si même variable; sinon colorer selon z
            if row.split(" : ")[0] == col.split(" : ")[0]:
                cell.fill = grey_fill
            else:
                z = tableau_z.loc[row, col]
                if z > 2:
                    cell.fill = green_fill
                elif z < -2:
                    cell.fill = red_fill

        # passer à la prochaine ligne
        excel_row += 1

        # ✅ Ligne vide après chaque bloc de variable
        next_row_var = toutes_modalites[ toutes_modalites.index(row) + 1].split(" : ")[0] if toutes_modalites.index(row) + 1 < len(toutes_modalites) else None
        if next_row_var != var_name:
            excel_row += 1  # ligne vide pour séparation

    wb.save(output_filename)
    print(f"✅ Fichier Excel généré : {output_filename}")

def tableaux_croises_khi2(df, variables, output_filename, max_modalites=10):
  result = tableaux_croises_multi(df, variables, max_modalites=max_modalites)
  couleurs_khi2(result, output_filename=output_filename)


def generate_excel(df, columns, qcm_columns=[], col_poids=None ,filename="resultats.xlsx", max_modalites=10, max_modalites_tri=30):
    dataframes = []

    # 📊 Tris à plat
    for column in columns:
        dataframes.append(tri_a_plat(df, column, col_poids=col_poids))
    for column in qcm_columns:
        dataframes.append(tri_a_plat(df, column, QCM=True, col_poids=col_poids))

    with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
        workbook = writer.book

        # -------- FEUILLE 1 : TRI À PLAT --------
        worksheet = workbook.add_worksheet("Tri à plat")
        row = 1
        for df_tri in dataframes:
            if len(df_tri) <= max_modalites_tri:
                question = df_tri.columns[0]
                worksheet.write(f"A{row}", question)
                df_tri = df_tri.rename(columns={df_tri.columns[0]: "Réponses"})
                df_tri.to_excel(writer, index=False, startrow=row, sheet_name="Tri à plat")
                row += len(df_tri) + 4

        # Formats %
        percentage_format = workbook.add_format({"num_format": "0%"})
        worksheet.set_column("C:C", None, percentage_format)
        if col_poids:
            worksheet.set_column("E:E", None, percentage_format)

        # -------- FEUILLE 2 : TABLEAUX KHI² --------
        resultats = tableaux_croises_multi(df, columns, max_modalites=max_modalites)
        worksheet2 = workbook.add_worksheet("Khi²")
        _write_khi2_to_worksheet(worksheet2, resultats, workbook)

    print(f"✅ Fichier Excel généré : {filename}")


def _write_khi2_to_worksheet(ws, resultats, workbook):
    """Écrit les tableaux Khi² dans une feuille xlsxwriter existante"""
    green_fill = workbook.add_format({"bg_color": "#C6EFCE"})
    red_fill = workbook.add_format({"bg_color": "#FFC7CE"})
    grey_fill = workbook.add_format({"bg_color": "#D9D9D9"})

    # Récupération de toutes les modalités
    toutes_modalites = []
    seen_modalites = set()
    for (var_row, var_col), (df_counts, _) in resultats.items():
        for m in df_counts.index:
            if m not in seen_modalites:
                toutes_modalites.append(m)
                seen_modalites.add(m)
        for m in df_counts.columns:
            if m not in seen_modalites:
                toutes_modalites.append(m)
                seen_modalites.add(m)

    tableau_perc = pd.DataFrame(index=toutes_modalites, columns=toutes_modalites, dtype=float)
    tableau_z = pd.DataFrame(index=toutes_modalites, columns=toutes_modalites, dtype=float)

    # Calcul résidus std
    for (var_row, var_col), (df_counts, df_perc) in resultats.items():
        observed = df_counts.to_numpy()
        try:
            chi2, p, dof, expected = chi2_contingency(observed)
            residus_std = (observed - expected) / np.sqrt(expected)
        except Exception:
            continue
        for i, row in enumerate(df_counts.index):
            for j, col in enumerate(df_counts.columns):
                tableau_perc.loc[row, col] = df_perc.loc[row, col]
                tableau_z.loc[row, col] = residus_std[i, j] if p < 0.05 else 0

    # Écriture Excel
    ws.write(0, 0, "Modalité")
    for j, col in enumerate(toutes_modalites):
        ws.write(0, j + 1, col)

    current_var = None
    excel_row = 1
    for row in toutes_modalites:
        var_name = row.split(" : ")[0]
        if current_var is None or var_name != current_var:
            ws.write(excel_row, 0, var_name)
            excel_row += 1
            current_var = var_name

        ws.write(excel_row, 0, row)
        for j, col in enumerate(toutes_modalites):
            val = tableau_perc.loc[row, col]
            if pd.isna(val):
                continue
            cell_val = f"{val:.2f}%"
            z = tableau_z.loc[row, col]

            if row.split(" : ")[0] == col.split(" : ")[0]:
                ws.write(excel_row, j + 1, cell_val, grey_fill)
            elif z > 2:
                ws.write(excel_row, j + 1, cell_val, green_fill)
            elif z < -2:
                ws.write(excel_row, j + 1, cell_val, red_fill)
            else:
                ws.write(excel_row, j + 1, cell_val)

        excel_row += 1
        # saut de ligne si changement de variable
        next_var = (
            toutes_modalites[toutes_modalites.index(row) + 1].split(" : ")[0]
            if toutes_modalites.index(row) + 1 < len(toutes_modalites)
            else None
        )
        if next_var != var_name:
            excel_row += 1


##############################################################################################################################################################


st.set_page_config(page_title="Analyse questionnaire", layout="wide")

st.title("📊 Analyse automatique de questionnaire")

# --- Upload du fichier ---
uploaded_file = st.file_uploader("Chargez votre fichier (CSV ou Excel)", type=["csv", "xlsx"])

if uploaded_file is not None:
    # Charger le fichier
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    st.success(f"✅ Fichier chargé : {uploaded_file.name}")
    st.write("Aperçu des données :", df.head())

    # --- Sélection colonnes ---
    colonnes = st.multiselect("Colonnes simples à analyser", df.columns.tolist())
    colonnes_qcm = st.multiselect("Colonnes QCM (choix multiples)", df.columns.tolist())
    col_poids = st.selectbox("Colonne de pondération (optionnel)", [""] + df.columns.tolist())

    # --- Nombre maximum de modalités pour tri à plat ---
    max_modalites_tri = st.number_input(
    "Nombre maximum de modalités pour tri à plat", min_value=2, value=30
    )

    # --- Nombre maximum de modalités pour tableaux croisés ---
    max_modalites = st.number_input(
        "Nombre maximum de modalités pour tableaux croisés", min_value=2, value=10
    )

    # --- Bouton lancer analyse ---
    if st.button("🚀 Lancer l'analyse"):
        fichier_sortie = "resultats.xlsx"
        generate_excel(
            df,
            columns=colonnes,
            qcm_columns=colonnes_qcm,
            col_poids=col_poids if col_poids else None,
            filename=fichier_sortie,
            max_modalites=max_modalites,
            max_modalites_tri=max_modalites_tri
        )
        st.success("✅ Analyse terminée !")

        # --- Téléchargement du fichier ---
        with open(fichier_sortie, "rb") as f:
            st.download_button(
                label="📥 Télécharger le fichier Excel",
                data=f,
                file_name=fichier_sortie,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
